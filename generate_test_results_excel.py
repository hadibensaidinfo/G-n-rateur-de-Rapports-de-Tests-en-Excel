import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import PatternFill


def ___GenereCasDeTest(path):
    """Cette fonction permet de retourner les résultats des tests sous forme d'un tableau de dictionnaires,
    où chaque dictionnaire contient le nom du cas de test, son statut, et le message d'erreur.
    ARG: 
    path(string): C'est le chemin du fichier xUnit."""
    tree = ET.parse(path)
    root = tree.findall('.//testcase')
    tab = []
    for child in root:
        dictionnaire = {}
        dictionnaire['path'] = child.attrib['classname']
        dictionnaire['name'] = child.attrib['name']
        failure = child.find('failure')
        if failure is not None:
            dictionnaire['message'] = failure.attrib.get('message', 'No message')
            dictionnaire['Status'] = 'Failed'
        else:
            dictionnaire['message'] = None
            dictionnaire['Status'] = 'Passed'
        tab.append(dictionnaire)
    return tab

def ___MettreLeResultatDansFichierExcel(path):
    dataTestCase = ___GenereCasDeTest(path)
    testPassed =0
    testFailed=0
    workbook  = Workbook()
    ws = workbook.active
    ws.title = 'Résultat'
    # Ajouter des données dans les titres (première ligne)
    ws['A1'] = 'Path de test'
    ws['B1'] = 'Cas De Test'
    ws['C1'] = 'Status'
    ws['D1'] = 'Message D\'erreur'

    # Définir un style pour la première ligne (titres)
    blue_fill = PatternFill(start_color="7F7FFF", end_color="7F7FFF", fill_type="solid")
    for cell in ws[1]:  # Première ligne
        cell.fill = blue_fill

    # Définir des styles pour les statuts (Passed: vert, Failed: rouge)
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Ajouter les lignes de données et colorier chaque ligne
    for  item in dataTestCase:
        # Ajouter une ligne avec les données
        row = [item['path'],item['name'], item['Status'], item['message']]
        ws.append(row)

        # Récupérer la ligne de la feuille après l'ajout
        last_row = ws.max_row 

        # Colorier la ligne selon le statut
        if item['Status'] == 'Passed':
            testPassed += 1 
            # Colorier la ligne en vert si Passed
            for col in range(1, 4):
                ws.cell(row=last_row, column=col).fill = green_fill
        else:
            testFailed += 1 
            # Colorier la ligne en rouge si Failed
            for col in range(1, 4):
                ws.cell(row=last_row, column=col).fill = red_fill
    ___CreeClasseurStatistics(workbook, testFailed,testPassed)
    workbook.save("resultatDeCasDeTest.xlsx")


def ___CreeClasseurStatistics(workbook, testFailed,testPassed):
    """Ajoute une feuille de statistiques au fichier Excel."""
    new_sheet = workbook.create_sheet(title="Statistiques")
    new_sheet['A1'] = 'Description'
    new_sheet['B1'] = 'Valeur'

    # Définir un style pour les titres
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for cell in new_sheet[1]:
        cell.fill = yellow_fill


    # Ajouter les données statistiques
    stats = [
        ['Nombre total de cas de test', testPassed + testFailed],
        ['Nombre de cas de test réussis', testPassed],
        ['Nombre de cas de test échoués', testFailed],
    ]
    for row in stats:
        new_sheet.append(row)



___MettreLeResultatDansFichierExcel('xUnit.xml')
